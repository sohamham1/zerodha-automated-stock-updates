import json
import os
import re
import sys
from html import unescape
from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


PALETTE = {
    "navy": "14263D",
    "steel": "355C7D",
    "mist": "EAF0F6",
    "sand": "ECE6DC",
    "cream": "F8F5EF",
    "green": "216E39",
    "green_soft": "DDF2E3",
    "red": "A63D40",
    "red_soft": "F8D8DA",
    "amber": "A46C1F",
    "amber_soft": "FBECC7",
    "ink": "1F2933",
    "muted": "667085",
    "white": "FFFFFF",
}

THIN_BORDER = Border(
    left=Side(style="thin", color=PALETTE["sand"]),
    right=Side(style="thin", color=PALETTE["sand"]),
    top=Side(style="thin", color=PALETTE["sand"]),
    bottom=Side(style="thin", color=PALETTE["sand"]),
)


def clean_display_text(value):
    text = unescape(str(value or ""))
    replacements = {
        "â€“": "-",
        "â€”": "-",
        "â€˜": "'",
        "â€™": "'",
        "â€œ": '"',
        "â€�": '"',
        "â€¢": "-",
        "â€¦": "...",
        "Â": "",
        "₹": "Rs ",
    }
    for bad, good in replacements.items():
        text = text.replace(bad, good)
    text = text.replace("\u200b", "").replace("\ufeff", "")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def clean_items(items):
    return [clean_display_text(item) for item in (items or []) if clean_display_text(item)]


def paragraph_text(value):
    return escape(clean_display_text(value))


def paragraph_lines(items):
    cleaned = clean_items(items)
    if not cleaned:
        return "None called out."
    return "<br/>".join([f"&bull; {escape(item)}" for item in cleaned])


def format_currency(value):
    return f"INR {float(value or 0):,.2f}"


def format_pct(value):
    return f"{float(value or 0):.2f}%"


def format_date_label(value):
    return clean_display_text(value) or "Date unavailable"


def holding_display_name(holding):
    return clean_display_text(holding.get("displayName") or holding.get("companyName") or holding.get("symbol") or "Unknown")


def holding_ticker(holding):
    return clean_display_text(holding.get("ticker") or holding.get("symbol") or "")


def holding_label(holding):
    return f"{holding_display_name(holding)} ({holding_ticker(holding)})"


def freshness_badge(item):
    freshness = clean_display_text(item.get("freshness", "")).lower()
    return "Older reference" if freshness in {"fallback", "stale"} else ""


def citation_meta(item):
    bits = [clean_display_text(item.get("source", "Unknown source"))]
    if item.get("publishedAtLabel"):
        bits.append(format_date_label(item.get("publishedAtLabel")))
    stale = freshness_badge(item)
    if stale:
        bits.append(stale)
    return " | ".join([bit for bit in bits if bit])


def sentiment_fill(sentiment):
    sentiment = clean_display_text(sentiment).lower()
    if sentiment == "bullish":
        return PatternFill("solid", fgColor=PALETTE["green_soft"])
    if sentiment == "cautious":
        return PatternFill("solid", fgColor=PALETTE["red_soft"])
    if sentiment == "neutral":
        return PatternFill("solid", fgColor=PALETTE["amber_soft"])
    return PatternFill("solid", fgColor=PALETTE["mist"])


def heat_fill(value):
    value = float(value or 0)
    if value >= 50:
        return PatternFill("solid", fgColor="CBEFD7")
    if value >= 15:
        return PatternFill("solid", fgColor="E5F6EA")
    if value >= 0:
        return PatternFill("solid", fgColor="F2F6E2")
    if value >= -15:
        return PatternFill("solid", fgColor="FCEFD3")
    if value >= -30:
        return PatternFill("solid", fgColor="F9D9D7")
    return PatternFill("solid", fgColor="F2C6CC")


def set_widths(sheet, widths):
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width


def style_header_row(sheet, row_idx, fill):
    for cell in sheet[row_idx]:
        cell.font = Font(name="Aptos", size=10, bold=True, color=PALETTE["white"])
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_body_range(sheet, start_row, end_row, left_cols=None):
    left_cols = left_cols or set()
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            cell.font = Font(name="Aptos", size=10, color=PALETTE["ink"])
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="left" if cell.column in left_cols else "center",
                vertical="center",
                wrap_text=True,
            )


def autosize_with_cap(sheet, min_width=8, max_width=42):
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[letter].width = max(min(length + 2, max_width), min_width)


def apply_cover(sheet, title, subtitle, generated_at):
    sheet.merge_cells("A1:J2")
    cell = sheet["A1"]
    cell.value = clean_display_text(title)
    cell.font = Font(name="Aptos Display", size=24, bold=True, color=PALETTE["white"])
    cell.fill = PatternFill("solid", fgColor=PALETTE["navy"])
    cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 32
    sheet.row_dimensions[2].height = 32

    sheet.merge_cells("A3:J4")
    sub = sheet["A3"]
    sub.value = clean_display_text(subtitle)
    sub.font = Font(name="Aptos", size=11, color=PALETTE["ink"])
    sub.fill = PatternFill("solid", fgColor=PALETTE["cream"])
    sub.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    sheet.merge_cells("H5:J5")
    meta = sheet["H5"]
    meta.value = f"Generated {clean_display_text(generated_at)[:10]}"
    meta.font = Font(name="Aptos", size=9, italic=True, color=PALETTE["muted"])
    meta.alignment = Alignment(horizontal="right")


def write_kpi_card(sheet, top_left, title, value, accent):
    row = sheet[top_left].row
    col = sheet[top_left].column
    top = sheet.cell(row=row, column=col)
    bottom = sheet.cell(row=row + 1, column=col)
    top.value = clean_display_text(title)
    bottom.value = clean_display_text(value)
    top.font = Font(name="Aptos", size=9, bold=True, color=PALETTE["muted"])
    bottom.font = Font(name="Aptos Display", size=16, bold=True, color=accent)
    top.fill = PatternFill("solid", fgColor=PALETTE["cream"])
    bottom.fill = PatternFill("solid", fgColor=PALETTE["white"])
    top.border = THIN_BORDER
    bottom.border = THIN_BORDER
    top.alignment = Alignment(horizontal="left")
    bottom.alignment = Alignment(horizontal="left")


def summarize_categories(holdings):
    counts = {}
    for holding in holdings:
        for item in holding.get("evidence", []):
            category = clean_display_text(item.get("category"))
            if category:
                counts[category] = counts.get(category, 0) + 1
    return sorted(counts.items(), key=lambda pair: pair[1], reverse=True)


def recommendation_bucket(holding):
    trends = holding.get("trends") or {}
    rsi = trends.get("rsi14")
    sma = trends.get("sma50")
    last_price = holding.get("lastPrice") or 0
    return_pct = holding.get("returnPct") or 0
    bc = holding.get("brokerageConsensus") or {}
    buy = bc.get("buy", 0)
    sell = bc.get("sell", 0)
    neutral = bc.get("neutral", 0)

    if return_pct <= -20 or sell > buy:
        return "review"
    if last_price and sma and last_price > sma and (rsi is None or 45 <= rsi <= 68) and buy >= sell:
        return "buy_watch"
    if return_pct >= 20 or neutral >= buy:
        return "neutral"
    return "neutral"


def executive_notes(report):
    holdings = report.get("holdings", [])
    if not holdings:
        return ["No equity holdings were available in the saved snapshot."]

    biggest = max(holdings, key=lambda item: item["portfolioWeight"])
    winners = sorted(holdings, key=lambda item: item["returnPct"], reverse=True)
    laggards = sorted(holdings, key=lambda item: item["returnPct"])
    buy_watch = [holding_label(item) for item in holdings if recommendation_bucket(item) == "buy_watch"]
    neutral_list = [holding_label(item) for item in holdings if recommendation_bucket(item) == "neutral"]
    review_list = [holding_label(item) for item in holdings if recommendation_bucket(item) == "review"]
    category_counts = summarize_categories(holdings)

    notes = [
        f"Portfolio concentration remains high in {holding_label(biggest)} at {format_pct(biggest['portfolioWeight'])}; this is the single name most likely to drive weekly portfolio direction.",
        f"Strongest all-time winner: {holding_label(winners[0])} at {format_pct(winners[0]['returnPct'])}.",
        f"Highest-priority review candidate: {holding_label(laggards[0])} is the deepest drag at {format_pct(laggards[0]['returnPct'])}.",
    ]

    if buy_watch:
        notes.append(f"Buy or watchlist names: {', '.join(buy_watch[:3])}.")
    if neutral_list:
        notes.append(f"Steadier hold names: {', '.join(neutral_list[:3])}.")
    if review_list:
        notes.append(f"Review or trim candidates: {', '.join(review_list[:3])}.")
    if category_counts:
        top_categories = ", ".join([name for name, _ in category_counts[:3]])
        notes.append(f"Recent evidence is dominated by {top_categories} updates, so the portfolio is reacting more to company news flow than to isolated noise.")
    if report.get("summary", {}).get("ordersCount", 0) == 0:
        notes.append("No recent order activity was captured in this run, so this brief should be read as a holdings-health update rather than a trading journal.")
    return notes[:6]


def add_dashboard(workbook, report):
    sheet = workbook.active
    sheet.title = "Dashboard"
    apply_cover(
        sheet,
        "Portfolio Weekly Intelligence Report",
        "A concise weekly brief grounded in Zerodha holdings, recent public evidence, broker coverage, and explicit data gaps.",
        report["generatedAt"],
    )

    summary = report["summary"]
    invested = summary.get("totalInvested") or sum(h["investedValue"] for h in report["holdings"])
    pnl = summary["totalPnl"]
    ret = 0 if not invested else (pnl / invested) * 100
    pnl_color = PALETTE["green"] if pnl >= 0 else PALETTE["red"]

    write_kpi_card(sheet, "A7", "Current Value", format_currency(summary["totalValue"]), PALETTE["navy"])
    write_kpi_card(sheet, "C7", "Invested Value", format_currency(invested), PALETTE["steel"])
    write_kpi_card(sheet, "E7", "All-time P&L", format_currency(pnl), pnl_color)
    write_kpi_card(sheet, "G7", "All-time Return", format_pct(ret), pnl_color)

    strongest = max(report["holdings"], key=lambda item: item["returnPct"])
    weakest = min(report["holdings"], key=lambda item: item["returnPct"])
    largest = max(report["holdings"], key=lambda item: item["portfolioWeight"])

    sheet["J7"] = "What matters now"
    sheet["J7"].font = Font(name="Aptos", size=11, bold=True, color=PALETTE["navy"])
    insight_cards = [
        ("Largest position", f"{holding_label(largest)} | {format_pct(largest['portfolioWeight'])}", PALETTE["steel"]),
        ("Strongest winner", f"{holding_label(strongest)} | {format_pct(strongest['returnPct'])}", PALETTE["green"]),
        ("Weakest return", f"{holding_label(weakest)} | {format_pct(weakest['returnPct'])}", PALETTE["red"]),
        (
            "Stance mix",
            f"{summary['sentimentCounts']['bullish']} bullish / {summary['sentimentCounts'].get('neutral', 0)} neutral / {summary['sentimentCounts']['cautious']} cautious",
            PALETTE["amber"],
        ),
    ]
    row = 8
    for title, value, color in insight_cards:
        sheet.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
        sheet.merge_cells(start_row=row + 1, start_column=10, end_row=row + 1, end_column=11)
        t = sheet.cell(row=row, column=10)
        v = sheet.cell(row=row + 1, column=10)
        t.value = clean_display_text(title)
        v.value = clean_display_text(value)
        t.font = Font(name="Aptos", size=9, bold=True, color=PALETTE["muted"])
        v.font = Font(name="Aptos Display", size=11, bold=True, color=color)
        t.fill = PatternFill("solid", fgColor=PALETTE["cream"])
        v.fill = PatternFill("solid", fgColor=PALETTE["white"])
        t.border = THIN_BORDER
        v.border = THIN_BORDER
        row += 3

    notes = report.get("summary", {}).get("notes") or executive_notes(report)
    sheet["A12"] = "Executive notes"
    sheet["A12"].font = Font(name="Aptos", size=11, bold=True, color=PALETTE["navy"])
    for idx, note in enumerate(notes, start=13):
        sheet.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=8)
        cell = sheet.cell(row=idx, column=1)
        cell.value = f"- {clean_display_text(note)}"
        cell.font = Font(name="Aptos", size=10, color=PALETTE["ink"])
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    start_row = 20
    sheet[f"A{start_row}"] = "Portfolio snapshot"
    sheet[f"A{start_row}"].font = Font(name="Aptos", size=11, bold=True, color=PALETTE["navy"])
    headers = ["Holding", "Weight %", "Current Value", "All-time P&L", "Return %", "Sentiment", "Broker Coverage (30d)", "Consensus Mix"]
    header_row = start_row + 1
    for i, header in enumerate(headers, start=1):
        sheet.cell(row=header_row, column=i, value=header)
    style_header_row(sheet, header_row, PALETTE["steel"])

    holdings = sorted(report["holdings"], key=lambda item: item["currentValue"], reverse=True)
    for offset, holding in enumerate(holdings, start=1):
        r = header_row + offset
        bc = holding.get("brokerageConsensus", {})
        values = [
            holding_label(holding),
            round(holding["portfolioWeight"], 2),
            format_currency(holding["currentValue"]),
            format_currency(holding["pnl"]),
            round(holding["returnPct"], 2),
            clean_display_text(holding["summary"]["sentiment"]).title(),
            bc.get("scannedCount", 0),
            f"B {bc.get('buy',0)} / N {bc.get('neutral', 0)} / S {bc.get('sell',0)}",
        ]
        for i, value in enumerate(values, start=1):
            sheet.cell(row=r, column=i, value=value)
        sheet.cell(row=r, column=5).fill = heat_fill(holding["returnPct"])
        sheet.cell(row=r, column=6).fill = sentiment_fill(holding["summary"]["sentiment"])
        sheet.cell(row=r, column=4).font = Font(
            name="Aptos",
            size=10,
            bold=True,
            color=PALETTE["green"] if holding["pnl"] >= 0 else PALETTE["red"],
        )

    end_row = header_row + len(holdings)
    style_body_range(sheet, header_row + 1, end_row, left_cols={1, 8})
    set_widths(sheet, {"A": 30, "B": 12, "C": 18, "D": 18, "E": 12, "F": 12, "G": 14, "H": 24, "I": 4, "J": 4})


def add_heatmap_sheet(workbook, report):
    sheet = workbook.create_sheet("Heatmap")
    apply_cover(
        sheet,
        "Portfolio Heatmap",
        "A quick visual scan of exposure, all-time return, P&L, sentiment, and broker coverage across the portfolio.",
        report["generatedAt"],
    )
    headers = ["Ticker", "Weight %", "Return %", "All-time P&L", "Broker Coverage (30d)", "Buy", "Neutral", "Sell", "Sentiment"]
    row = 7
    for i, header in enumerate(headers, start=1):
        sheet.cell(row=row, column=i, value=header)
    style_header_row(sheet, row, PALETTE["steel"])

    holdings = sorted(report["holdings"], key=lambda item: item["portfolioWeight"], reverse=True)
    for offset, holding in enumerate(holdings, start=1):
        r = row + offset
        bc = holding.get("brokerageConsensus", {})
        values = [
            holding_ticker(holding),
            round(holding["portfolioWeight"], 2),
            round(holding["returnPct"], 2),
            format_currency(holding["pnl"]),
            bc.get("scannedCount", 0),
            bc.get("buy", 0),
            bc.get("neutral", 0),
            bc.get("sell", 0),
            clean_display_text(holding["summary"]["sentiment"]).title(),
        ]
        for i, value in enumerate(values, start=1):
            sheet.cell(row=r, column=i, value=value)
        sheet.cell(row=r, column=2).fill = heat_fill(holding["portfolioWeight"] - 20)
        sheet.cell(row=r, column=3).fill = heat_fill(holding["returnPct"])
        sheet.cell(row=r, column=4).fill = heat_fill(holding["returnPct"])
        sheet.cell(row=r, column=9).fill = sentiment_fill(holding["summary"]["sentiment"])
        for c in (5, 6, 7, 8):
            sheet.cell(row=r, column=c).fill = PatternFill("solid", fgColor=PALETTE["cream"])

    style_body_range(sheet, row + 1, row + len(holdings), left_cols={1, 9})
    set_widths(sheet, {"A": 16, "B": 12, "C": 12, "D": 16, "E": 14, "F": 8, "G": 8, "H": 8, "I": 14})
    sheet.freeze_panes = "A8"


def add_holdings_sheet(workbook, report):
    sheet = workbook.create_sheet("Holdings")
    apply_cover(
        sheet,
        "Holdings Detail",
        "Ground-truth holding values directly from the Zerodha snapshot, with cleaned display names and narrative metadata.",
        report["generatedAt"],
    )
    headers = [
        "Exchange", "Ticker", "Display Name", "ISIN", "Qty", "Buy Avg", "Last Price",
        "Invested Value", "Current Value", "All-time P&L", "Return %", "Weight %",
        "Weekly Move %", "50-day SMA", "14-day RSI", "Volatility %", "Sentiment", "Broker Coverage (30d)",
    ]
    row = 7
    for i, header in enumerate(headers, start=1):
        sheet.cell(row=row, column=i, value=header)
    style_header_row(sheet, row, PALETTE["navy"])

    for offset, holding in enumerate(report["holdings"], start=1):
        r = row + offset
        trends = holding.get("trends") or {}
        sma50 = trends.get("sma50")
        rsi14 = trends.get("rsi14")
        vol = trends.get("volatility")
        values = [
            clean_display_text(holding.get("exchangeLabel") or holding["exchange"]),
            holding_ticker(holding),
            holding_display_name(holding),
            clean_display_text(holding.get("isin") or ""),
            holding["quantity"],
            holding["averagePrice"],
            holding["lastPrice"],
            holding["investedValue"],
            holding["currentValue"],
            holding["pnl"],
            holding["returnPct"],
            holding["portfolioWeight"],
            holding["weeklyChangePct"],
            format_currency(sma50) if sma50 is not None else "N/A",
            f"{rsi14:.2f}" if rsi14 is not None else "N/A",
            f"{vol:.2f}%" if vol is not None else "N/A",
            clean_display_text(holding["summary"]["sentiment"]).title(),
            holding.get("brokerageConsensus", {}).get("scannedCount", 0),
        ]
        for i, value in enumerate(values, start=1):
            sheet.cell(row=r, column=i, value=value)
        sheet.cell(row=r, column=10).font = Font(
            name="Aptos",
            size=10,
            bold=True,
            color=PALETTE["green"] if holding["pnl"] >= 0 else PALETTE["red"],
        )
        sheet.cell(row=r, column=11).fill = heat_fill(holding["returnPct"])
        sheet.cell(row=r, column=17).fill = sentiment_fill(holding["summary"]["sentiment"])

    style_body_range(sheet, row + 1, row + len(report["holdings"]), left_cols={1, 2, 3, 4, 17})
    set_widths(sheet, {"A": 10, "B": 14, "C": 28, "D": 18, "E": 8, "F": 12, "G": 12, "H": 14, "I": 14, "J": 14, "K": 12, "L": 12, "M": 12, "N": 14, "O": 12, "P": 14, "Q": 12, "R": 14})
    sheet.freeze_panes = "A8"


def add_stock_summary_sheet(workbook, report):
    sheet = workbook.create_sheet("Stock Summaries")
    apply_cover(
        sheet,
        "Narrative Summary",
        "Cleaned portfolio narratives for what matters now, what moved this week, confidence, and explicit evidence gaps.",
        report["generatedAt"],
    )
    headers = ["Holding", "Sentiment", "Confidence", "What Matters Now", "What Moved This Week", "Risk / Watchpoints", "Top Evidence"]
    row = 7
    for i, header in enumerate(headers, start=1):
        sheet.cell(row=row, column=i, value=header)
    style_header_row(sheet, row, PALETTE["steel"])

    for offset, holding in enumerate(report["holdings"], start=1):
        r = row + offset
        citations = [
            f"{clean_display_text(item.get('title'))} ({citation_meta(item)})"
            for item in (holding.get("summary", {}).get("citations") or [])
        ]
        values = [
            holding_label(holding),
            clean_display_text(holding["summary"]["sentiment"]).title(),
            f"{clean_display_text(holding['summary']['confidence']).title()} - {clean_display_text(holding['summary'].get('confidenceReason', ''))}",
            clean_display_text(holding["summary"]["rationale"]),
            clean_display_text(holding["summary"]["whyMoving"]),
            "\n".join([f"- {item}" for item in clean_items(holding["summary"]["watchpoints"] + holding["summary"].get("missingEvidence", []))]),
            "\n".join([f"- {item}" for item in citations]) or "- No linked evidence",
        ]
        for i, value in enumerate(values, start=1):
            sheet.cell(row=r, column=i, value=value)
        sheet.cell(row=r, column=2).fill = sentiment_fill(holding["summary"]["sentiment"])
        line_count = max(str(value).count("\n") + 1 for value in values[2:])
        sheet.row_dimensions[r].height = max(36, min(18 + line_count * 14, 150))

    style_body_range(sheet, row + 1, row + len(report["holdings"]), left_cols={1, 3, 4, 5, 6, 7})
    set_widths(sheet, {"A": 28, "B": 12, "C": 34, "D": 36, "E": 36, "F": 36, "G": 42})
    sheet.freeze_panes = "A8"


def safe_get_float(data, key, default=0.0):
    try:
        value = data.get(key)
        if value is None:
            return default
        return float(value)
    except Exception:
        return default


def add_account_transactions_sheet(workbook, report):
    sheet = workbook.create_sheet("Account & Transactions")
    apply_cover(
        sheet,
        "Account Margins & Orders History",
        "Available cash margins and recent order history tracked from Zerodha.",
        report["generatedAt"],
    )

    sheet["A7"] = "Available Margins / Cash"
    sheet["A7"].font = Font(name="Aptos", size=12, bold=True, color=PALETTE["navy"])

    margins = report.get("margins") or {}
    equity = margins.get("equity", {}) or margins
    headers_margins = ["Segment", "Available Cash", "Used Margin", "Available Margin"]
    row_idx = 9
    for i, header in enumerate(headers_margins, start=1):
        sheet.cell(row=row_idx, column=i, value=header)
    style_header_row(sheet, row_idx, PALETTE["steel"])

    row_idx += 1
    sheet.cell(row=row_idx, column=1, value="Equity")
    sheet.cell(row=row_idx, column=2, value=format_currency(safe_get_float(equity, "cash") or safe_get_float(margins, "cash") or 0.0))
    sheet.cell(row=row_idx, column=3, value=format_currency(safe_get_float(equity, "used") or safe_get_float(margins, "used") or 0.0))
    sheet.cell(row=row_idx, column=4, value=format_currency(safe_get_float(equity, "net") or safe_get_float(margins, "net") or 0.0))

    commodity = margins.get("commodity", {})
    if commodity:
        row_idx += 1
        sheet.cell(row=row_idx, column=1, value="Commodity")
        sheet.cell(row=row_idx, column=2, value=format_currency(safe_get_float(commodity, "cash") or 0.0))
        sheet.cell(row=row_idx, column=3, value=format_currency(safe_get_float(commodity, "used") or 0.0))
        sheet.cell(row=row_idx, column=4, value=format_currency(safe_get_float(commodity, "net") or 0.0))

    style_body_range(sheet, 9, row_idx, left_cols={1})

    row_idx += 3
    sheet.cell(row=row_idx, column=1, value="Recent Orders History").font = Font(name="Aptos", size=12, bold=True, color=PALETTE["navy"])
    headers_orders = ["Time", "Symbol", "Exchange", "Transaction", "Order Type", "Quantity", "Price", "Status", "Status Message"]
    row_idx += 1
    order_header_row = row_idx
    for i, header in enumerate(headers_orders, start=1):
        sheet.cell(row=row_idx, column=i, value=header)
    style_header_row(sheet, row_idx, PALETTE["navy"])

    for order in report.get("orders") or []:
        row_idx += 1
        sheet.cell(row=row_idx, column=1, value=clean_display_text(order.get("order_timestamp", order.get("timestamp", ""))))
        sheet.cell(row=row_idx, column=2, value=clean_display_text(order.get("tradingsymbol", order.get("symbol", ""))))
        sheet.cell(row=row_idx, column=3, value=clean_display_text(order.get("exchange", "")))
        sheet.cell(row=row_idx, column=4, value=clean_display_text(order.get("transaction_type", "")))
        sheet.cell(row=row_idx, column=5, value=clean_display_text(order.get("order_type", "")))
        sheet.cell(row=row_idx, column=6, value=order.get("quantity", 0))
        sheet.cell(row=row_idx, column=7, value=format_currency(order.get("price", 0.0)))
        sheet.cell(row=row_idx, column=8, value=clean_display_text(order.get("status", "")))
        sheet.cell(row=row_idx, column=9, value=clean_display_text(order.get("status_message", "")))

    if row_idx > order_header_row:
        style_body_range(sheet, order_header_row + 1, row_idx, left_cols={1, 2, 8, 9})

    set_widths(sheet, {"A": 22, "B": 14, "C": 12, "D": 14, "E": 14, "F": 10, "G": 16, "H": 14, "I": 24})


def add_sources_sheet(workbook, report):
    sheet = workbook.create_sheet("Sources")
    apply_cover(
        sheet,
        "Evidence & Sources",
        "Full evidence inventory and brokerage links reviewed for this portfolio brief.",
        report["generatedAt"],
    )
    headers = ["Ticker", "Display Name", "Category", "Title", "Source", "Published At", "Freshness", "URL"]
    row = 7
    for i, header in enumerate(headers, start=1):
        sheet.cell(row=row, column=i, value=header)
    style_header_row(sheet, row, PALETTE["steel"])

    current_row = row + 1
    for holding in report["holdings"]:
        for item in holding.get("evidence", []):
            values = [
                holding_ticker(holding),
                holding_display_name(holding),
                clean_display_text(item.get("category", "")),
                clean_display_text(item.get("title", "")),
                clean_display_text(item.get("source", "")),
                clean_display_text(item.get("publishedAtLabel") or item.get("publishedAt") or ""),
                clean_display_text(item.get("freshness", "")),
                clean_display_text(item.get("url", "")),
            ]
            for i, value in enumerate(values, start=1):
                cell = sheet.cell(row=current_row, column=i, value=value)
                if i == 8 and value:
                    cell.hyperlink = str(value)
                    cell.style = "Hyperlink"
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            current_row += 1

        for item in holding.get("brokerageConsensus", {}).get("items", []):
            values = [
                holding_ticker(holding),
                holding_display_name(holding),
                "brokerage_consensus",
                clean_display_text(item.get("title", "")),
                clean_display_text(item.get("broker") or item.get("source", "")),
                clean_display_text(item.get("publishedAtLabel") or item.get("publishedAt") or ""),
                clean_display_text(item.get("freshness", "")),
                clean_display_text(item.get("url", "")),
            ]
            for i, value in enumerate(values, start=1):
                cell = sheet.cell(row=current_row, column=i, value=value)
                if i == 8 and value:
                    cell.hyperlink = str(value)
                    cell.style = "Hyperlink"
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            current_row += 1

    if current_row > row + 1:
        style_body_range(sheet, row + 1, current_row - 1, left_cols={1, 2, 3, 4, 5, 6, 7, 8})
        for row_idx in range(row + 1, current_row):
            sheet.cell(row=row_idx, column=8).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    set_widths(sheet, {"A": 12, "B": 24, "C": 18, "D": 44, "E": 22, "F": 16, "G": 12, "H": 50})
    sheet.freeze_panes = "A8"


def build_workbook(report, output_dir):
    workbook = Workbook()
    add_dashboard(workbook, report)
    add_heatmap_sheet(workbook, report)
    add_holdings_sheet(workbook, report)
    add_stock_summary_sheet(workbook, report)
    add_account_transactions_sheet(workbook, report)
    add_sources_sheet(workbook, report)
    for sheet in workbook.worksheets:
        autosize_with_cap(sheet)

    profile = report.get("profile", "default")
    date_str = report.get("generatedAt", "")[:10]
    file_path = os.path.join(output_dir, f"report_{date_str}_{profile}.xlsx")
    workbook.save(file_path)
    return file_path


def pdf_styles():
    styles = getSampleStyleSheet()
    return {
        "title": ParagraphStyle("Title", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=26, leading=30, textColor=colors.HexColor("#14263D"), alignment=TA_LEFT, spaceAfter=8),
        "heading": ParagraphStyle("Heading", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=13.5, leading=17, textColor=colors.HexColor("#14263D"), spaceBefore=8, spaceAfter=5),
        "body": ParagraphStyle("Body", parent=styles["BodyText"], fontName="Helvetica", fontSize=9.8, leading=13.2, textColor=colors.HexColor("#1F2933"), spaceAfter=0),
        "muted": ParagraphStyle("Muted", parent=styles["BodyText"], fontName="Helvetica", fontSize=8.3, leading=10.5, textColor=colors.HexColor("#667085"), spaceAfter=0),
        "hero_value": ParagraphStyle("HeroValue", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=16.5, leading=18.5, textColor=colors.white, spaceAfter=2),
        "hero_label": ParagraphStyle("HeroLabel", parent=styles["BodyText"], fontName="Helvetica", fontSize=8.5, leading=10, textColor=colors.HexColor("#D7E4F1"), spaceAfter=0),
        "hero_metric": ParagraphStyle("HeroMetric", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=8.2, leading=10, textColor=colors.HexColor("#A9C2DD"), spaceAfter=2),
        "hero_subtitle": ParagraphStyle("HeroSubtitle", parent=styles["BodyText"], fontName="Helvetica", fontSize=10.8, leading=15, textColor=colors.HexColor("#DDE8F5"), spaceAfter=0),
        "hero_meta": ParagraphStyle("HeroMeta", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=9.6, leading=12, textColor=colors.white, spaceAfter=0, alignment=TA_LEFT),
        "section_label": ParagraphStyle("SectionLabel", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=8.4, leading=10, textColor=colors.HexColor("#355C7D"), spaceAfter=3),
        "card_body": ParagraphStyle("CardBody", parent=styles["BodyText"], fontName="Helvetica", fontSize=9.3, leading=12.3, textColor=colors.HexColor("#1F2933"), spaceAfter=0),
        "stock_title": ParagraphStyle("StockTitle", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=14.5, leading=18, textColor=colors.HexColor("#14263D"), spaceAfter=1),
        "stock_meta": ParagraphStyle("StockMeta", parent=styles["BodyText"], fontName="Helvetica", fontSize=8.9, leading=11.3, textColor=colors.HexColor("#667085"), spaceAfter=0),
        "source_link": ParagraphStyle("SourceLink", parent=styles["BodyText"], fontName="Helvetica", fontSize=8.2, leading=10.4, textColor=colors.HexColor("#1F2933"), spaceAfter=0),
    }


def mini_kpi_table(report, styles):
    invested = report["summary"].get("totalInvested") or sum(h["investedValue"] for h in report["holdings"])
    pnl = report["summary"]["totalPnl"]
    ret = 0 if not invested else (pnl / invested) * 100
    data = [[
        Paragraph("<b>Invested value</b><br/>" + paragraph_text(format_currency(invested)), styles["body"]),
        Paragraph("<b>Current value</b><br/>" + paragraph_text(format_currency(report["summary"]["totalValue"])), styles["body"]),
        Paragraph("<b>All-time P&amp;L</b><br/>" + paragraph_text(format_currency(pnl)), styles["body"]),
        Paragraph("<b>All-time return</b><br/>" + paragraph_text(format_pct(ret)), styles["body"]),
    ]]
    table = Table(data, colWidths=[4.15 * cm] * 4)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8F5EF")),
        ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#ECE6DC")),
        ("INNERGRID", (0, 0), (-1, -1), 0.7, colors.HexColor("#ECE6DC")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    return table


def cover_insight_table(report):
    strongest = max(report["holdings"], key=lambda item: item["returnPct"])
    weakest = min(report["holdings"], key=lambda item: item["returnPct"])
    largest = max(report["holdings"], key=lambda item: item["portfolioWeight"])
    data = [
        ["Largest position", f"{holding_label(largest)} at {format_pct(largest['portfolioWeight'])} of portfolio value"],
        ["Best all-time return", f"{holding_label(strongest)} at {format_pct(strongest['returnPct'])}"],
        ["Weakest all-time return", f"{holding_label(weakest)} at {format_pct(weakest['returnPct'])}"],
    ]
    table = Table(data, colWidths=[4.3 * cm, 11.0 * cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#ECE6DC")),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#ECE6DC")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#355C7D")),
    ]))
    return table


def top_holdings_table(report):
    holdings = sorted(report["holdings"], key=lambda item: item["portfolioWeight"], reverse=True)
    rows = [[
        Paragraph("Holding", pdf_styles()["body"]),
        Paragraph("Weight %", pdf_styles()["body"]),
        Paragraph("Return %", pdf_styles()["body"]),
        Paragraph("Sentiment", pdf_styles()["body"]),
    ]]
    for holding in holdings:
        rows.append([
            Paragraph(paragraph_text(holding_label(holding)), pdf_styles()["body"]),
            Paragraph(paragraph_text(format_pct(holding["portfolioWeight"])), pdf_styles()["body"]),
            Paragraph(paragraph_text(format_pct(holding["returnPct"])), pdf_styles()["body"]),
            Paragraph(paragraph_text(clean_display_text(holding["summary"]["sentiment"]).title()), pdf_styles()["body"]),
        ])
    table = Table(rows, colWidths=[7.0 * cm, 2.6 * cm, 2.7 * cm, 4.3 * cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#14263D")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#ECE6DC")),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#ECE6DC")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("WORDWRAP", (0, 0), (-1, -1), "CJK"),
    ]))
    return table


def executive_summary_points(report):
    holdings = report["holdings"]
    biggest = max(holdings, key=lambda item: item["portfolioWeight"])
    strongest = max(holdings, key=lambda item: item["returnPct"])
    weakest = min(holdings, key=lambda item: item["returnPct"])
    summary = report["summary"]["sentimentCounts"]
    invested = report["summary"].get("totalInvested") or sum(h["investedValue"] for h in holdings)
    total_return = 0 if not invested else (report["summary"]["totalPnl"] / invested) * 100
    return [
        f"The portfolio is up {format_currency(report['summary']['totalPnl'])} all-time, with an all-time return of {format_pct(total_return)}.",
        f"{holding_label(biggest)} is now the dominant position at {format_pct(biggest['portfolioWeight'])} of portfolio value, so it will drive portfolio direction most visibly.",
        f"{holding_label(strongest)} remains the strongest winner at {format_pct(strongest['returnPct'])}, while {holding_label(weakest)} is the weakest all-time return at {format_pct(weakest['returnPct'])}.",
        f"The current stance mix reads {summary['bullish']} bullish, {summary.get('neutral', 0)} neutral, and {summary['cautious']} cautious holdings.",
    ]


def action_points(report):
    holdings = report["holdings"]
    largest = max(holdings, key=lambda item: item["portfolioWeight"])
    review_candidates = sorted(
        [item for item in holdings if recommendation_bucket(item) == "review"],
        key=lambda item: item["portfolioWeight"],
        reverse=True,
    )
    neutral_candidates = [item for item in holdings if recommendation_bucket(item) == "neutral"]
    buy_candidates = [item for item in holdings if recommendation_bucket(item) == "buy_watch"]
    weakest = min(holdings, key=lambda item: item["returnPct"])
    points = [
        f"Anchor attention on {holding_label(largest)} first because it carries the largest portfolio weight at {format_pct(largest['portfolioWeight'])}.",
        f"Revisit the original thesis for {holding_label(weakest)} because it has the deepest all-time drawdown at {format_pct(weakest['returnPct'])}.",
    ]
    if buy_candidates:
        points.append(f"Buy or watchlist names from the latest snapshot: {', '.join(holding_label(item) for item in buy_candidates[:3])}.")
    if neutral_candidates:
        points.append(f"Steadier hold names: {', '.join(holding_label(item) for item in neutral_candidates[:3])}.")
    if review_candidates:
        points.append(f"Review or trim candidates: {', '.join(holding_label(item) for item in review_candidates[:3])}.")
    points.append("Treat broker-consensus counts carefully when coverage is limited; several names in this report still have incomplete outside research visibility.")
    return points


def sentiment_color(sentiment):
    sentiment = clean_display_text(sentiment).lower()
    if sentiment == "bullish":
        return colors.HexColor("#216E39")
    if sentiment == "cautious":
        return colors.HexColor("#A63D40")
    return colors.HexColor("#A46C1F")


def hero_banner(report, styles):
    invested = report["summary"].get("totalInvested") or sum(h["investedValue"] for h in report["holdings"])
    pnl = report["summary"]["totalPnl"]
    ret = 0 if not invested else (pnl / invested) * 100

    def metric_block(label, value):
        return Paragraph(
            f"<font color='#A9C2DD'><b>{paragraph_text(label)}</b></font><br/><font color='#FFFFFF'><b>{paragraph_text(value)}</b></font>",
            styles["hero_meta"],
        )

    data = [
        [Paragraph("Weekly Portfolio Brief", styles["title"]), ""],
        [
            Paragraph(
                "A concise weekly read of portfolio health, concentration risk, recent evidence, and the clearest next actions.",
                styles["hero_subtitle"],
            ),
            Paragraph(
                f"<font color='#A9C2DD'>Generated</font><br/><font color='#FFFFFF'><b>{paragraph_text(report['generatedAt'][:10])}</b></font>",
                styles["hero_meta"],
            ),
        ],
        [Paragraph("AT A GLANCE", styles["hero_metric"]), Paragraph("KEY METRICS", styles["hero_metric"])],
        [metric_block("Current portfolio value", format_currency(report["summary"]["totalValue"])), metric_block("All-time portfolio P&L", format_currency(pnl))],
        [metric_block("All-time portfolio return", format_pct(ret)), metric_block("Equity holdings covered", str(report["summary"]["holdingsCount"]))],
    ]
    table = Table(data, colWidths=[11.6 * cm, 5.0 * cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#14263D")),
        ("SPAN", (0, 0), (1, 0)),
        ("LINEBELOW", (0, 1), (1, 1), 0.5, colors.HexColor("#355C7D")),
        ("LINEABOVE", (0, 2), (1, 2), 0.5, colors.HexColor("#355C7D")),
        ("LINEABOVE", (0, 4), (1, 4), 0.5, colors.HexColor("#355C7D")),
        ("LEFTPADDING", (0, 0), (-1, -1), 14),
        ("RIGHTPADDING", (0, 0), (-1, -1), 14),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 1), (1, 1), "RIGHT"),
        ("ALIGN", (1, 2), (1, 4), "RIGHT"),
    ]))
    return table


def insight_cards(title, items, styles):
    rows = [[Paragraph(paragraph_text(title), styles["heading"])]]
    for item in items:
        rows.append([Paragraph(f"&bull; {paragraph_text(item)}", styles["body"])])
    table = Table(rows, colWidths=[16.6 * cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F8F5EF")),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#ECE6DC")),
        ("INNERGRID", (0, 1), (-1, -1), 0.3, colors.HexColor("#F1ECE4")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    return table


def stock_header_table(holding, styles):
    sentiment = clean_display_text(holding["summary"]["sentiment"]).title()
    meta_left = f"{clean_display_text(holding.get('exchangeLabel') or holding['exchange'])} | Qty {holding['quantity']} | Buy avg {format_currency(holding['averagePrice'])} | Current {format_currency(holding['lastPrice'])}"
    meta_right = f"Confidence {clean_display_text(holding['summary']['confidence']).title()} | Weight {format_pct(holding['portfolioWeight'])}"
    data = [
        [Paragraph(paragraph_text(holding_label(holding)), styles["stock_title"]), Paragraph(f"<b>{paragraph_text(sentiment)}</b>", styles["stock_title"])],
        [Paragraph(paragraph_text(meta_left), styles["stock_meta"]), Paragraph(paragraph_text(meta_right), styles["stock_meta"])],
    ]
    table = Table(data, colWidths=[11.6 * cm, 5.0 * cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8F5EF")),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#ECE6DC")),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("ALIGN", (1, 0), (1, 1), "RIGHT"),
        ("TEXTCOLOR", (1, 0), (1, 0), sentiment_color(sentiment)),
    ]))
    return table


def stock_card_table(holding):
    bc = holding.get("brokerageConsensus", {})
    left = [
        ["Original buy price", format_currency(holding["averagePrice"])],
        ["Current price", format_currency(holding["lastPrice"])],
        ["Current value", format_currency(holding["currentValue"])],
        ["All-time return", format_pct(holding["returnPct"])],
    ]
    right = [
        ["All-time P&L", format_currency(holding["pnl"])],
        ["Portfolio weight", format_pct(holding["portfolioWeight"])],
        ["Weekly move", format_pct(holding["weeklyChangePct"])],
        ["Broker coverage (30d)", str(bc.get("scannedCount", 0))],
    ]
    rows = [["Metric", "Value", "Metric", "Value"]] + [[a, b, c, d] for (a, b), (c, d) in zip(left, right)]
    table = Table(rows, colWidths=[3.6 * cm, 4.0 * cm, 3.6 * cm, 3.8 * cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#14263D")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#ECE6DC")),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#ECE6DC")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    return table


def narrative_box(label, text, styles, fill):
    table = Table(
        [[Paragraph(paragraph_text(label.upper()), styles["section_label"])], [Paragraph(text, styles["card_body"])]],
        colWidths=[8.05 * cm],
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(f"#{fill}")),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#ECE6DC")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    return table


def dual_narrative_row(left_label, left_text, right_label, right_text, styles):
    table = Table(
        [[
            narrative_box(left_label, left_text, styles, PALETTE["white"]),
            narrative_box(right_label, right_text, styles, PALETTE["cream"]),
        ]],
        colWidths=[8.1 * cm, 8.1 * cm],
    )
    table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    return table


def evidence_table(holding, styles):
    rows = [["Top evidence", "Why it matters"]]
    for item in (holding.get("summary", {}).get("citations") or [])[:3]:
        title = paragraph_text(item.get("title", "Untitled source"))
        url = clean_display_text(item.get("url", ""))
        meta = paragraph_text(citation_meta(item))
        source_cell = Paragraph(
            f"<link href='{escape(url)}' color='#1D4ED8'><u>{title}</u></link>",
            styles["source_link"],
        ) if url else Paragraph(title, styles["source_link"])
        meta_cell = Paragraph(meta, styles["muted"])
        rows.append([source_cell, meta_cell])
    if len(rows) == 1:
        rows.append([Paragraph("No linked evidence captured.", styles["source_link"]), Paragraph("Evidence quality is thin for this holding.", styles["muted"])])
    table = Table(rows, colWidths=[11.3 * cm, 4.9 * cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#14263D")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#ECE6DC")),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#F1ECE4")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    return table


def build_pdf(report, output_dir):
    profile = report.get("profile", "default")
    date_str = report.get("generatedAt", "")[:10]
    file_path = os.path.join(output_dir, f"report_{date_str}_{profile}.pdf")
    styles = pdf_styles()
    doc = SimpleDocTemplate(
        file_path,
        pagesize=A4,
        rightMargin=1.4 * cm,
        leftMargin=1.4 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )

    story = [
        hero_banner(report, styles),
        Spacer(1, 0.3 * cm),
        insight_cards("Key weekly takeaways", executive_summary_points(report), styles),
        Spacer(1, 0.22 * cm),
        insight_cards("Top actions", action_points(report), styles),
        Spacer(1, 0.22 * cm),
        Paragraph("Portfolio shape", styles["heading"]),
        cover_insight_table(report),
        Spacer(1, 0.22 * cm),
        mini_kpi_table(report, styles),
        Spacer(1, 0.22 * cm),
        Paragraph("Portfolio weight snapshot", styles["heading"]),
        top_holdings_table(report),
    ]
    for note in report["summary"].get("notes", []):
        story.append(Paragraph(paragraph_text(note), styles["muted"]))

    story.append(PageBreak())

    for index, holding in enumerate(report["holdings"]):
        confidence_text = clean_display_text(holding["summary"].get("confidenceReason", ""))
        missing_evidence = holding["summary"].get("missingEvidence") or []
        evidence_gap_text = confidence_text
        if missing_evidence:
            evidence_gap_text = f"{paragraph_text(confidence_text)}<br/><br/>{paragraph_lines(missing_evidence)}" if confidence_text else paragraph_lines(missing_evidence)
        else:
            evidence_gap_text = paragraph_text(confidence_text or "No major evidence gaps were flagged.")

        story.extend([
            stock_header_table(holding, styles),
            Spacer(1, 0.18 * cm),
            stock_card_table(holding),
            Spacer(1, 0.18 * cm),
            dual_narrative_row(
                "What matters now",
                paragraph_text(holding["summary"]["rationale"]),
                "What moved this week",
                paragraph_text(holding["summary"]["whyMoving"]),
                styles,
            ),
            Spacer(1, 0.18 * cm),
            dual_narrative_row(
                "Risk / watchpoints",
                paragraph_lines(holding["summary"]["watchpoints"]),
                "Confidence / evidence gaps",
                evidence_gap_text,
                styles,
            ),
            Spacer(1, 0.18 * cm),
            Paragraph("Top evidence", styles["heading"]),
            evidence_table(holding, styles),
        ])
        if index != len(report["holdings"]) - 1:
            story.append(PageBreak())

    doc.build(story)
    return file_path


def main():
    report_json_path = Path(sys.argv[1])
    output_dir = Path(sys.argv[2])
    include_pdf = sys.argv[3] == "1"
    report = json.loads(report_json_path.read_text(encoding="utf-8"))
    build_workbook(report, str(output_dir))
    if include_pdf:
        build_pdf(report, str(output_dir))
    print("ok")


if __name__ == "__main__":
    main()
